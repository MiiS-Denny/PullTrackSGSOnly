# -*- coding: utf-8 -*-
"""
拉力值紀錄SGS_Only
- 含登入驗證（PBKDF2-SHA256）
- 上傳 Excel 範本，輸入 12 筆資料，下載 <原檔名>-out.xlsx
- 自動維持原本公式與兩張圖表的資料範圍
- 同日多筆：需勾選確認後，自動於日期加 -01 / -02 / ...
- 日期輸入可為 YYYY/MM/DD[-LL]、YYYYMMDDLL、YYYY/MM/DD、YYYYMMDD

使用方式：
    pip install -r requirements.txt
    streamlit run app.py
"""

import hmac, hashlib, binascii, re, tempfile
from datetime import datetime
from io import BytesIO
from collections import defaultdict

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from openpyxl.chart import Reference


# =========================
# 帳密（PBKDF2-SHA256）
# =========================
PWD_DB = {
    "Charles": {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "32ae892164a22af5f83261bd239ed304", "hash": "27fb5fb7bbe2629d8c53dbbdf021423cdb4e7015e5858deafb3a0e405139bb40"},
    "Hsiang":  {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "9c31cdf98b82aa1741154680e456e3e0", "hash": "292e30442d243ea5f82879f1ce71f9ff2dc600f7234a075ba3ee130f45eb29b4"},
    "Sandy":   {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "dd8fbb2a735b076e5cff3bdee67fc3cf", "hash": "7bff0a1388c1447e934552175786d2fa5b9bc9b17ac3d9da246182dd7ec31e35"},
    "Min":     {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "a4a89474d39a1d89ac652a56ccd33301", "hash": "7d788d76be27923209c08aba44fdfc0ca6ce5530ed4b91283810fd0c34bc1a0f"},
    "May":     {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "88d33f6eb3d9a6506b705c3810e7be0b", "hash": "53765f6d56af8c2e49f917c89d60212ab8aeec28d215c9e53cf394e897782631"},
    "Ping":    {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "4af5ee4403ad13cb6a2b0836da5d02b1", "hash": "1c1757b927959d2ef8897467f1c823753ec166f0d5c0a1a8ed5d91a84f2ab00d"},
    "Denny":   {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "bc88ba930b619a25dcce81e6ee616305", "hash": "3dfe81a7dd31acaf2816604c000637f328049d1ca9f13940e217ec51f3a5e7c7"},
    "Davina":  {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "8ce1cb7106316a21db1b48534d7d1833", "hash": "3a79b1feaa96cd7dc7dbced0bc2226d84da22ecda5a38d7d44a58f98e8c24b96"},
    "Arthur":  {"algo": "pbkdf2_sha256", "iter": 200000, "salt": "8e9a0b3e6c6dd1dccd6964101b5af752", "hash": "0409292dedb20de507c7fae67d25f502998c80cb4fcace6758d8fedc042d5570"},
}

def _pbkdf2_sha256(password: str, salt_hex: str, iter: int, dklen: int = 32) -> str:
    salt = binascii.unhexlify(salt_hex)
    dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, iter, dklen=dklen)
    return binascii.hexlify(dk).decode('ascii')

def verify(username: str, password: str) -> bool:
    rec = PWD_DB.get(username)
    if not rec or rec.get('algo') != 'pbkdf2_sha256':
        return False
    expect_hex = rec['hash']
    got_hex = _pbkdf2_sha256(password, rec['salt'], rec['iter'])
    return hmac.compare_digest(expect_hex, got_hex)


# =========================
# Excel 欄位定義與工具
# =========================
COL_DATE = 1
COL_V_START = 2
COL_V_END = 7
COL_XBAR = 8
COL_R = 9
COL_CL_XBAR = 10
COL_UCL_XBAR = 11
COL_LCL_XBAR = 12
COL_CL_R = 13
COL_UCL_R = 14
COL_LCL_R = 15
COL_OWNER = 16

def find_last_data_row(ws, col=COL_DATE):
    max_row = ws.max_row
    for r in range(max_row, 1, -1):
        v = ws.cell(row=r, column=col).value
        if v not in (None, ""):
            return r
    return 1

def to_float_or_raise(s: str, name: str) -> float:
    try:
        return float(s)
    except Exception:
        raise ValueError(f"{name} 需為數字，收到：{s!r}")

def copy_cell_style(src, dst):
    if src.has_style:
        if src.font:        dst.font = Font(**src.font.__dict__)
        if src.alignment:   dst.alignment = Alignment(**src.alignment.__dict__)
        if src.border:      dst.border = Border(**src.border.__dict__)
        if src.fill:        dst.fill = PatternFill(**src.fill.__dict__)
        if src.protection:  dst.protection = Protection(**src.protection.__dict__)
        dst.number_format = src.number_format

def copy_row_styles(ws, from_row: int, to_row: int, col_start: int, col_end: int):
    ws.row_dimensions[to_row].height = ws.row_dimensions[from_row].height
    for c in range(col_start, col_end + 1):
        copy_cell_style(ws.cell(row=from_row, column=c), ws.cell(row=to_row, column=c))

# ---- 解析 / 正規化日期 ----
# 接受：
#  1) YYYY/MM/DD-LL
#  2) YYYYMMDDLL
#  3) YYYY/MM/DD        -> 無 LL
#  4) YYYYMMDD          -> 無 LL
def parse_date_input(s: str):
    s = (s or "").strip()
    if not s:
        return None, None  # 無輸入
    # 含 LL
    m = re.fullmatch(r"(\d{4})[\/\-]?(\d{1,2})[\/\-]?(\d{1,2})[\/\-](\d{1,2})", s)
    if not m:
        m = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(\d{2})", s)
    if m:
        y, mo, d, ll = map(int, m.groups())
        base = f"{y:04d}/{mo:02d}/{d:02d}"
        return base, f"{ll:02d}"

    # 無 LL
    m2 = re.fullmatch(r"(\d{4})[\/\.-]?(\d{1,2})[\/\.-]?(\d{1,2})", s)
    if m2:
        y, mo, d = map(int, m2.groups())
        base = f"{y:04d}/{mo:02d}/{d:02d}"
        return base, None

    raise ValueError(f"日期需為 YYYY/MM/DD-LL、YYYYMMDDLL、YYYY/MM/DD 或 YYYYMMDD，收到：{s!r}")

def compose_date(base: str, ll: int | str | None) -> str:
    if ll is None:
        ll = 0
    if isinstance(ll, str):
        try:
            ll = int(ll)
        except Exception:
            ll = 0
    return f"{base}-{ll:02d}"

# 從工作表掃描已存在的 base -> 已使用最大 LL
def scan_existing_ll(ws):
    last = find_last_data_row(ws, col=COL_DATE)
    used = defaultdict(int)  # base -> max_ll
    for r in range(2, last + 1):
        v = ws.cell(row=r, column=COL_DATE).value
        if v in (None, ""):
            continue
        if isinstance(v, datetime):
            base = v.strftime("%Y/%m/%d")
            ll = 0
        else:
            s = str(v).strip()
            m = re.fullmatch(r"(\d{4})[\/\-]?(\d{1,2})[\/\-]?(\d{1,2})[\/\-](\d{1,2})", s)
            if m:
                y, mo, d, ll = map(int, m.groups())
                base = f"{y:04d}/{mo:02d}/{d:02d}"
            else:
                m2 = re.fullmatch(r"(\d{4})[\/\.-]?(\d{1,2})[\/\.-]?(\d{1,2})", s)
                if m2:
                    y, mo, d = map(int, m2.groups())
                    base = f"{y:04d}/{mo:02d}/{d:02d}"
                    ll = 0
                else:
                    continue
        used[base] = max(used[base], int(ll))
    return used

def refresh_existing_two_charts(ws, last_row):
    charts = getattr(ws, "_charts", [])
    if not charts:
        return

    def _reset_chart_series_to_cols(chart, cols):
        anchor = getattr(chart, "anchor", None)
        title = getattr(chart, "title", None)
        y_title = getattr(getattr(chart, "y_axis", None), "title", None)

        chart.series = []
        cats = Reference(ws, min_col=COL_DATE, min_row=2, max_row=last_row)
        for col in cols:
            data = Reference(ws, min_col=col, min_row=1, max_row=last_row)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.title = title
        if hasattr(chart, "y_axis") and y_title is not None:
            chart.y_axis.title = y_title
        if anchor:
            chart.anchor = anchor

    _reset_chart_series_to_cols(charts[0], (COL_XBAR, COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR))
    if len(charts) >= 2:
        _reset_chart_series_to_cols(charts[1], (COL_R, COL_CL_R, COL_UCL_R, COL_LCL_R))

def _select_sheet(wb, preferred_name: str):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name], preferred_name
    candidates = [n for n in wb.sheetnames if getattr(wb[n], "sheet_state", "visible") == "visible"]
    used = candidates[0] if candidates else wb.sheetnames[0]
    return wb[used], used

def _append_one(ws, date_norm: str, values: list[float], work_order_text: str):
    last_row = find_last_data_row(ws, col=COL_DATE)
    new_row = last_row + 1

    copy_row_styles(ws, from_row=last_row, to_row=new_row, col_start=COL_DATE, col_end=COL_LCL_R)

    ws.cell(row=new_row, column=COL_DATE).value = date_norm

    for i, col in enumerate(range(COL_V_START, COL_V_END + 1), start=1):
        ws.cell(row=new_row, column=col).value = values[i-1]

    v_start = ws.cell(row=new_row, column=COL_V_START).coordinate
    v_end   = ws.cell(row=new_row, column=COL_V_END).coordinate
    ws.cell(row=new_row, column=COL_XBAR).value = f"=AVERAGE({v_start}:{v_end})"
    ws.cell(row=new_row, column=COL_R).value    = f"=MAX({v_start}:{v_end})-MIN({v_start}:{v_end})"

    if last_row >= 2:
        for col in (COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR, COL_CL_R, COL_UCL_R, COL_LCL_R):
            ws.cell(row=new_row, column=col).value = ws.cell(row=last_row, column=col).value

    pcell = ws.cell(row=new_row, column=COL_OWNER)
    pcell.value = work_order_text or ""
    pcell.font = Font(name="Calibri", size=11)
    pcell.alignment = Alignment(wrap_text=True)

    return new_row

def append_many_from_template_bytes(template_bytes: bytes, rows_to_add: list, sheet_name: str = "Data"):
    """回傳：(out_bytes, used_sheet)"""
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tf.write(template_bytes)
        tf.flush()
        wb = load_workbook(tf.name, data_only=False)
    finally:
        tf.close()

    ws, used_sheet = _select_sheet(wb, sheet_name)

    # 先掃描既有日期的最大層號
    existing_max_ll = scan_existing_ll(ws)

    last_row_after = find_last_data_row(ws, col=COL_DATE)
    appended = False

    for idx, row in enumerate(rows_to_add, start=1):
        raw_date = (row.get("date") or "").strip()
        work_order = (row.get("owner") or "").strip()
        vals = row.get("values") or []

        if not raw_date:
            continue

        base, ll_str = parse_date_input(raw_date)  # ll_str 可能為 None
        if len(vals) != 6:
            wb.close()
            raise ValueError(f"第 {idx} 筆：Value_1~Value_6 必須 6 個數字")
        nums = [to_float_or_raise(v, f"第 {idx} 筆 Value_{i+1}") for i, v in enumerate(vals)]

        # 若未指定 LL，就用（既有最大 + 目前批次同日累加）的下一號
        if ll_str is None:
            existing_max_ll[base] += 1
            ll_use = existing_max_ll[base]
        else:
            # 若指定了 LL，也更新既有最大（避免下一筆繼續同一天）
            ll_use = int(ll_str)
            existing_max_ll[base] = max(existing_max_ll[base], ll_use)

        date_norm = compose_date(base, ll_use)

        last_row_after = _append_one(ws, date_norm, nums, work_order)
        appended = True

    if not appended:
        wb.close()
        raise ValueError("沒有可新增的資料：12 筆輸入裡的日期全為空白")

    refresh_existing_two_charts(ws, last_row_after)

    bio = BytesIO()
    wb.save(bio)
    wb.close()
    bio.seek(0)
    return bio.read(), used_sheet

def read_last_date_str_from_template_bytes(template_bytes: bytes, sheet_name: str = "Data") -> str:
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tf.write(template_bytes)
        tf.flush()
        wb = load_workbook(tf.name, data_only=True, read_only=True)
    finally:
        tf.close()

    ws, _ = _select_sheet(wb, sheet_name)
    r = find_last_data_row(ws, col=COL_DATE)
    if r <= 1:
        wb.close()
        return ""
    v = ws.cell(row=r, column=COL_DATE).value
    wb.close()
    if v is None or v == "":
        return ""

    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d-00")

    s = str(v).strip()
    m = re.fullmatch(r"(\d{4})[\/\-]?(\d{1,2})[\/\-]?(\d{1,2})[\/\-](\d{1,2})", s)
    if m:
        y, mo, d, ll = m.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}-{int(ll):02d}"
    m2 = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(\d{2})", s)
    if m2:
        y, mo, d, ll = m2.groups()
        return f"{y}/{mo}/{d}-{ll}"
    m3 = re.fullmatch(r"(\d{4})[\/\.-]?(\d{1,2})[\/\.-]?(\d{1,2})", s)
    if m3:
        y, mo, d = m3.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}-00"
    return s


# =========================
# Streamlit App
# =========================
st.set_page_config(page_title="拉力值紀錄SGS_Only", layout="wide")

if "authed_user" not in st.session_state:
    st.session_state.authed_user = None

def login_form():
    st.title("登入 / Login")
    with st.form("login"):
        col1, col2 = st.columns(2)
        with col1:
            username = st.selectbox("帳號 (Username)", options=sorted(PWD_DB.keys()))
        with col2:
            password = st.text_input("密碼 (Password)", type="password")
        submitted = st.form_submit_button("登入")
    if submitted:
        if verify(username, password):
            st.session_state.authed_user = username
            st.success(f"歡迎，{username}")
            st.rerun()
        else:
            st.error("帳號或密碼錯誤")
    st.stop()

if st.session_state.authed_user is None:
    login_form()

st.sidebar.write(f"已登入：**{st.session_state.authed_user}**")
st.title("拉力值紀錄SGS_Only")

uploaded_file = st.file_uploader("上傳 Excel 範本 (.xlsx/.xlsm)", type=["xlsx", "xlsm"])
sheet_name = st.text_input("工作表名稱（找不到會退回第一個可見表）", value="Data")

if uploaded_file is None:
    st.info("請先上傳範本檔案。")
    st.stop()

template_bytes = uploaded_file.read()

# 顯示最後一筆日期
try:
    last_date = read_last_date_str_from_template_bytes(template_bytes, sheet_name=sheet_name)
    st.caption(f"最後一筆日期：{last_date or '—'}")
except Exception as e:
    st.warning(f"讀取最後一筆日期失敗：{e}")

# 準備 12 列空白輸入
rows = []
for _ in range(12):
    rows.append({
        "Date (YYYY/MM/DD 或 YYYY/MM/DD-LL 皆可)": "",
        "Value_1 (P1-1)": "",
        "Value_2 (P1-2)": "",
        "Value_3 (P1-3)": "",
        "Value_4 (P2-1)": "",
        "Value_5 (P2-2)": "",
        "Value_6 (P2-3)": "",
        "WO No.（工單號碼）": "",
    })
df = pd.DataFrame(rows)

st.write("請在下表輸入 12 筆資料（空白日期列會自動略過；若同日多筆會自動編號 -01、-02…）：")
edited = st.data_editor(df, num_rows="fixed", use_container_width=True, key="input_table")

# 檢查同日多筆（含與既有資料衝突時的『同日新筆』）
need_confirm_multi = False
dup_info = defaultdict(int)

for _, row in edited.iterrows():
    raw_date = str(row["Date (YYYY/MM/DD 或 YYYY/MM/DD-LL 皆可)"]).strip()
    if not raw_date or raw_date.lower() == "nan":
        continue
    try:
        base, ll = parse_date_input(raw_date)
        if ll is None:
            dup_info[base] += 1
            if dup_info[base] >= 2:
                need_confirm_multi = True
        else:
            key = (base, ll)
            dup_info[key] += 1
            if dup_info[key] >= 2:
                need_confirm_multi = True
    except Exception:
        pass

if need_confirm_multi:
    st.warning("偵測到**同一天多筆**或**同日重複層號**的輸入。若要繼續，請勾選下方確認，系統會自動以 -01、-02… 編號。")
    confirmed = st.checkbox("我確認同一天需要新增多筆資料，並同意自動編號（-01、-02…）", value=False, key="confirm_multi")
else:
    confirmed = True  # 沒有重複就不需要另外確認

if st.button("新增紀錄並下載新檔"):
    if not confirmed:
        st.error("尚未勾選確認同日多筆新增。")
        st.stop()

    rows_to_add = []
    for _, row in edited.iterrows():
        raw_date = str(row["Date (YYYY/MM/DD 或 YYYY/MM/DD-LL 皆可)"]).strip()
        if not raw_date or raw_date.lower() == "nan":
            continue
        rows_to_add.append({
            "date": raw_date,
            "values": [
                str(row["Value_1 (P1-1)"]).strip(),
                str(row["Value_2 (P1-2)"]).strip(),
                str(row["Value_3 (P1-3)"]).strip(),
                str(row["Value_4 (P2-1)"]).strip(),
                str(row["Value_5 (P2-2)"]).strip(),
                str(row["Value_6 (P2-3)"]).strip(),
            ],
            "owner": str(row["WO No.（工單號碼）"]).strip(),
        })
    try:
        out_bytes, used_sheet = append_many_from_template_bytes(template_bytes, rows_to_add, sheet_name=sheet_name)
        base = uploaded_file.name.rsplit(".", 1)[0]
        out_name = f"{base}-out.xlsx"
        st.success(f"已完成（Sheet: {used_sheet}）。")
        st.download_button(
            "下載處理後檔案",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(str(e))
