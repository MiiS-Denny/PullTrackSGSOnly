import re, os, tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from openpyxl.chart import Reference

# 欄位定義（與原工具一致）
COL_DATE = 1
COL_V_START = 2
COL_V_END = 7
COL_XBAR = 8
COL_R = 9
COL_CL_XBAR = 10
COL_UCL_XBAR = 11
COL_LCL_XBAR = 12
COL_CL_R = 13
COL_UCL_R = 14
COL_LCL_R = 15
COL_OWNER = 16

def find_last_data_row(ws, col=COL_DATE):
    max_row = ws.max_row
    for r in range(max_row, 1, -1):
        v = ws.cell(row=r, column=col).value
        if v not in (None, ""):
            return r
    return 1

def to_float_or_raise(s: str, name: str) -> float:
    try:
        return float(s)
    except Exception:
        raise ValueError(f"{name} 需為數字，收到：{s!r}")

def copy_cell_style(src, dst):
    if src.has_style:
        if src.font:        dst.font = Font(**src.font.__dict__)
        if src.alignment:   dst.alignment = Alignment(**src.alignment.__dict__)
        if src.border:      dst.border = Border(**src.border.__dict__)
        if src.fill:        dst.fill = PatternFill(**src.fill.__dict__)
        if src.protection:  dst.protection = Protection(**src.protection.__dict__)
        dst.number_format = src.number_format

def copy_row_styles(ws, from_row: int, to_row: int, col_start: int, col_end: int):
    ws.row_dimensions[to_row].height = ws.row_dimensions[from_row].height
    for c in range(col_start, col_end + 1):
        copy_cell_style(ws.cell(row=from_row, column=c), ws.cell(row=to_row, column=c))

def normalize_date_ll_input(s: str) -> str:
    s = (s or "").strip()
    m = re.fullmatch(r"(\d{4})[/-]?(\d{2})[/-]?(\d{2})[-/]?(\d{2})", s)
    if not m:
        raise ValueError(f"日期需為 YYYY/MM/DD-LL 或 YYYYMMDDLL，收到：{s!r}")
    y, mo, d, ll = m.groups()
    return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}-{int(ll):02d}"

def _reset_chart_series_to_cols(ws, chart, cols, last_row):
    anchor = getattr(chart, "anchor", None)
    title = getattr(chart, "title", None)
    y_title = getattr(getattr(chart, "y_axis", None), "title", None)

    chart.series = []
    cats = Reference(ws, min_col=COL_DATE, min_row=2, max_row=last_row)

    for col in cols:
        data = Reference(ws, min_col=col, min_row=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.title = title
    if hasattr(chart, "y_axis") and y_title is not None:
        chart.y_axis.title = y_title
    if anchor:
        chart.anchor = anchor

def refresh_existing_two_charts(ws, last_row):
    charts = getattr(ws, "_charts", [])
    if not charts:
        return
    _reset_chart_series_to_cols(
        ws, charts[0],
        (COL_XBAR, COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR),
        last_row
    )
    if len(charts) >= 2:
        _reset_chart_series_to_cols(
            ws, charts[1],
            (COL_R, COL_CL_R, COL_UCL_R, COL_LCL_R),
            last_row
        )

def _select_sheet(wb, preferred_name: str):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name], preferred_name
    candidates = [n for n in wb.sheetnames if getattr(wb[n], "sheet_state", "visible") == "visible"]
    used = candidates[0] if candidates else wb.sheetnames[0]
    return wb[used], used

def _append_one(ws, date_norm: str, values: list[float], work_order_text: str):
    last_row = find_last_data_row(ws, col=COL_DATE)
    new_row = last_row + 1

    copy_row_styles(ws, from_row=last_row, to_row=new_row, col_start=COL_DATE, col_end=COL_LCL_R)

    ws.cell(row=new_row, column=COL_DATE).value = date_norm

    for i, col in enumerate(range(COL_V_START, COL_V_END + 1), start=1):
        ws.cell(row=new_row, column=col).value = values[i-1]

    v_start = ws.cell(row=new_row, column=COL_V_START).coordinate
    v_end   = ws.cell(row=new_row, column=COL_V_END).coordinate
    ws.cell(row=new_row, column=COL_XBAR).value = f"=AVERAGE({v_start}:{v_end})"
    ws.cell(row=new_row, column=COL_R).value    = f"=MAX({v_start}:{v_end})-MIN({v_start}:{v_end})"

    if last_row >= 2:
        for col in (COL_CL_XBAR, COL_UCL_XBAR, COL_LCL_XBAR, COL_CL_R, COL_UCL_R, COL_LCL_R):
            ws.cell(row=new_row, column=col).value = ws.cell(row=last_row, column=col).value

    pcell = ws.cell(row=new_row, column=COL_OWNER)
    pcell.value = work_order_text or ""
    pcell.font = Font(name="Calibri", size=11)
    pcell.alignment = Alignment(wrap_text=True)

    return new_row

def append_many(template_bytes: bytes, rows_to_add: list, sheet_name: str = "Data"):
    # 以 NamedTemporaryFile 方式載入 bytes（openpyxl 需要檔案路徑）
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tf.write(template_bytes)
        tf.flush()
        wb = load_workbook(tf.name, data_only=False)
    finally:
        tf.close()

    ws, used_sheet = _select_sheet(wb, sheet_name)

    last_row_after = find_last_data_row(ws, col=COL_DATE)
    appended = False

    for idx, row in enumerate(rows_to_add, start=1):
        raw_date = (row.get("date") or "").strip()
        work_order = (row.get("owner") or "").strip()
        vals = row.get("values") or []

        if not raw_date:
            continue

        date_norm = normalize_date_ll_input(raw_date)

        if len(vals) != 6:
            wb.close()
            raise ValueError(f"第 {idx} 筆：Value_1~Value_6 必須 6 個數字")
        nums = [to_float_or_raise(v, f"第 {idx} 筆 Value_{i+1}") for i, v in enumerate(vals)]

        last_row_after = _append_one(ws, date_norm, nums, work_order)
        appended = True

    if not appended:
        wb.close()
        raise ValueError("沒有可新增的資料：12 筆輸入裡的日期全為空白")

    refresh_existing_two_charts(ws, last_row_after)

    # 以記憶體輸出（回傳 bytes 與實際使用的 sheet 名稱）
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    wb.close()
    bio.seek(0)
    return bio.read(), used_sheet

def read_last_date_str_from_template_bytes(template_bytes: bytes, sheet_name: str = "Data") -> str:
    tf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        tf.write(template_bytes)
        tf.flush()
        wb = load_workbook(tf.name, data_only=True, read_only=True)
    finally:
        tf.close()

    ws, _ = _select_sheet(wb, sheet_name)
    r = find_last_data_row(ws, col=COL_DATE)
    if r <= 1:
        wb.close()
        return ""
    v = ws.cell(row=r, column=COL_DATE).value
    wb.close()
    if v is None or v == "":
        return ""

    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d-00")

    s = str(v).strip()
    m = re.fullmatch(r"(\d{4})[/-]?(\d{1,2})[/-]?(\d{1,2})[-/](\d{1,2})", s)
    if m:
        y, mo, d, ll = m.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}-{int(ll):02d}"
    m2 = re.fullmatch(r"(\d{4})(\d{2})(\d{2})(\d{2})", s)
    if m2:
        y, mo, d, ll = m2.groups()
        return f"{y}/{mo}/{d}-{ll}"
    m3 = re.fullmatch(r"(\d{4})[-/\.]?(\d{1,2})[-/\.]?(\d{1,2})", s)
    if m3:
        y, mo, d = m3.groups()
        return f"{int(y):04d}/{int(mo):02d}/{int(d):02d}-00"
    return s
